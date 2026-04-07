import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import uuid
import os

FILE = "bookings.xlsx"
HEADERS = ["id", "date", "time_start", "time_end", "user_id", "username", "comment"]


def init_file():
    """Создаёт файл если не существует"""
    if not os.path.exists(FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Бронирования"
        ws.append(HEADERS)
        wb.save(FILE)


def get_all_records():
    init_file()
    wb = load_workbook(FILE)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) <= 1:
        return []
    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:]]


def add_booking(date, time_start, time_end, user_id, username, comment=""):
    init_file()
    booking_id = str(uuid.uuid4())[:8].upper()
    wb = load_workbook(FILE)
    ws = wb.active
    ws.append([booking_id, date, time_start, time_end, str(user_id), username, comment])
    wb.save(FILE)
    return booking_id


def get_user_bookings(user_id):
    return [r for r in get_all_records() if str(r["user_id"]) == str(user_id)]


def get_bookings_by_date(date):
    return [r for r in get_all_records() if r["date"] == date]


def cancel_booking(booking_id, user_id):
    init_file()
    wb = load_workbook(FILE)
    ws = wb.active
    rows = list(ws.iter_rows())
    for i, row in enumerate(rows[1:], start=2):  # пропускаем заголовок
        if str(row[0].value) == booking_id and str(row[4].value) == str(user_id):
            ws.delete_rows(i)
            wb.save(FILE)
            return True
    return False


def is_time_available(date, time_start, time_end):
    bookings = get_bookings_by_date(date)
    new_start = datetime.strptime(time_start, "%H:%M")
    new_end = datetime.strptime(time_end, "%H:%M")
    for b in bookings:
        b_start = datetime.strptime(b["time_start"], "%H:%M")
        b_end = datetime.strptime(b["time_end"], "%H:%M")
        if not (new_end <= b_start or new_start >= b_end):
            return False
    return True
